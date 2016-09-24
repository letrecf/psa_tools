import codecs
from openpyxl import load_workbook
from collections import defaultdict
from . import helpers as h

symbols = [
    "circle", "diamond"
    ]
models = [
    "undefined", "repairable", "tested", "probability", "mission_time",
    "frequency", "non-repairable"
    ]
states = [
    "normal", "true", "false"
    ]
partypes = [
    "q", "fr", "f", "mttr", "ti", "ttf", "tm",
    "beta", "gamma", "delta", "alpha1", "alpha2", "alpha3", "alpha4"
    ]
ccf_models = [
    "beta", "MGL", "alpha"
    ]
pardist = [
    "none", "lognormal", "beta", "gamma", "normal",
    "uniform", "loguniform", "discrete"
    ]


class reader(object):

    def __init__(self, rsapath, xlsxpath):
        self.rsapath = rsapath
        self.xlsxpath = xlsxpath
        self._hasXLSX = True if xlsxpath else False
        self.encoding = "utf-8"
        self._use_cache = False
        self._fault_trees = []
        self._basic_events = []
        self._gates = []
        self._house_events = []
        self._ccf_groups = []
        self._attributes = []
        self._systems = []
        self._memo = []
        self._bc_sets = []
        self._analyses = []

    def getCache(self):
        return self._use_cache

    def setCache(self, val):
        self._use_cache = val
        if not(val):
            self._fault_trees = []
            self._basic_events = []
            self._gates = []
            self._house_events = []
            self._ccf_groups = []
            self._attributes = []
            self._systems = []
            self._memo = []
            self._bc_sets = []
            self._analyses = []

    use_cache = property(
        fget=getCache, fset=setCache,
        doc="flag for result caching"
        )

    def __read_section(self, section_name):
        section_length = len(section_name)
        with codecs.open(self.rsapath, "r", encoding=self.encoding) as fi:
            for line in fi:
                if line[:section_length] == section_name:
                    break
            for line in fi:
                if line[:2] == "//":
                    break
                else:
                    yield line

    def getFT(self):
        if self._use_cache and self._fault_trees:
            for i in self._fault_trees:
                yield i
            else:
                o = {}
                isMultilineText = False
                isInit = False
                for i in self.__read_section("//RECORD Fault Tree"):
                    d = i.split("\t")
                    if d[0] == "FTR":
                        if isInit:
                            if self._use_cache:
                                self._fault_trees.append(o)
                            yield o
                        d = i.split("\t")
                        o["name"] = h.get_quotedString(d[1])
                        if d[2] and (d[2].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[2]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[2])
                            else:
                                o["text"] = (d[2].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                if isInit:
                    yield o

    FaultTrees = property(
        fget=getFT,
        doc="Fault trees via iterator"
        )

    def getBE(self):
        if self._use_cache and self._basic_events:
            for i in self._basic_events:
                yield i
            else:
                if self._hasXLSX:
                    wb = load_workbook(filename=self.xlsxpath, read_only=True)
                    ws = wb["Basic Event"]
                    be_ini = dict(
                        (e[0].value, str(e[3].value).lower())
                        for e
                        in ws.rows[1:]
                        )
                else:
                    be_ini = {}
                o = {}
                isInit = False
                isMultilineText = False
                for i in self.__read_section("//RECORD Basic Event"):
                    d = i.split("\t")
                    if d[0] == "BEV":
                        if isInit:
                            if self._use_cache:
                                self._basic_events.append(o)
                            yield o
                        o["name"] = h.get_quotedString(d[1])
                        o["symbol"] = symbols[int(d[2]) - 1]
                        o["model"] = models[int(d[3])]
                        o["state"] = states[int(d[4])]
                        if d[5] and (d[5].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[5]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[5])
                            else:
                                o["text"] = (d[5].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["attributes"] = []
                        o["exchanges"] = []
                        o["parameters"] = {}
                        o["notes"] = []
                        o["initiator_enabler"] = \
                            be_ini.get(o["name"], "**unknown**")
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                        elif d[1] == "ATT":
                            if len(d) > 2:
                                o["attributes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["attributes"].append("**unknown**")
                        elif d[1] == "EXC":
                            if len(d) > 3:
                                o["exchanges"].append(
                                    (h.get_quotedString(d[2]),
                                    h.get_quotedString(d[3]))
                                    )
                            else:
                                o["exchanges"].append(
                                    (h.get_quotedString(d[2]),
                                    "**unknown**")
                                    )
                        elif d[1][0] == "P":
                            t = partypes[int(d[1][1:]) - 1]
                            if len(d) > 2:
                                o["parameters"][t] = \
                                    h.get_quotedString(d[2])
                            else:
                                o["parameters"][t] = "**unknown**"
                # emit last
                if isInit:
                    yield o

    BasicEvents = property(
        fget=getBE,
        doc="Basic events via iterator"
        )

    def getGates(self):
        if self._use_cache and self._gates:
            for i in self._gates:
                yield i
            else:
                o = {}
                isInit = False
                isMultilineText = False
                for i in self.__read_section("//RECORD Gate"):
                    d = i.split("\t")
                    if d[0] == "GAT":
                        if isInit:
                            if self._use_cache:
                                self._gates.append(o)
                            yield o
                        missing_page = True if d[4].isdigit() else False
                        o["name"] = h.get_quotedString(d[1])
                        o["type"] = d[2].strip()
                        o["state"] = states[int(d[3])]
                        if not(missing_page):
                            o["ft_page"] = h.get_quotedString(d[4])
                        else:
                            o["ft_page"] = "**missing**"
                        idx = 5 if not(missing_page) else 4
                        o["is_top"] = True if int(d[idx]) == 1 else False
                        idx += 1
                        if d[idx] and (d[idx].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[idx]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[idx])
                            else:
                                o["text"] = (d[idx].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["attributes"] = []
                        o["exchanges"] = []
                        o["inputs"] = []
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                        elif d[1] == "ATT":
                            if len(d) > 2:
                                o["attributes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["attributes"].append("**unknown**")
                        elif d[1] == "EXC":
                            if len(d) > 3:
                                o["exchanges"].append(
                                    (h.get_quotedString(d[2]),
                                    h.get_quotedString(d[3]))
                                    )
                            else:
                                o["exchanges"].append(
                                    (h.get_quotedString(d[2]),
                                    "**unknown**")
                                    )
                        elif d[1] == "GIN":
                            t = "**unknown**"
                            if d[2] == "GAT":
                                t = "gate"
                            elif d[2] == "BEV":
                                t = "basic_event"
                            elif d[2] == "HEV":
                                t = "house_event"
                            o["inputs"].append(
                                {"type": t,
                                "name": h.get_quotedString(d[3]),
                                "negated": True if int(d[4]) == -1 else False}
                                )
                # emit last
                if isInit:
                    yield o

    Gates = property(
        fget=getGates,
        doc="Gates via iterator"
        )

    def getHE(self):
        if self._use_cache and self._house_events:
            for i in self._house_events:
                yield i
            else:
                o = {}
                isMultilineText = False
                isInit = False
                for i in self.__read_section("//RECORD House event"):
                    d = i.split("\t")
                    if d[0] == "HEV":
                        if isInit:
                            if self._use_cache:
                                self._house_events.append(o)
                            yield o
                        d = i.split("\t")
                        o["name"] = h.get_quotedString(d[1])
                        o["state"] = states[int(d[2])]
                        if d[3] and (d[3].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[3]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[3])
                            else:
                                o["text"] = (d[3].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                if isInit:
                    yield o

    HouseEvents = property(
        fget=getHE,
        doc="House events via iterator"
        )

    def getCCF(self):
        if self._use_cache and self._ccf_groups:
            for i in self._ccf_groups:
                yield i
            else:
                o = {}
                isInit = False
                isMultilineText = False
                for i in self.__read_section("//RECORD CCF Group"):
                    d = i.split("\t")
                    if d[0] == "CCG":
                        if isInit:
                            if self._use_cache:
                                self._ccf_groups.append(o)
                            yield o
                        o["name"] = h.get_quotedString(d[1])
                        o["model"] = ccf_models[int(d[2]) - 1]
                        if d[3] and (d[3].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[3]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[3])
                            else:
                                o["text"] = (d[3].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["events"] = []
                        o["parameters"] = {}
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                        elif d[1] == "BEV":
                            if len(d) > 2:
                                o["events"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["events"].append("**unknown**")
                        elif d[1][0] == "P":
                            t = partypes[int(d[1][1:]) - 1]
                            if len(d) > 2:
                                o["parameters"][t] = \
                                    h.get_quotedString(d[2])
                            else:
                                o["parameters"][t] = "**unknown**"
                # emit last
                if isInit:
                    yield o

    CCF = property(
        fget=getCCF,
        doc="CCF groups via iterator"
        )

    def getAtt(self):
        if self._use_cache and self._attributes:
            for i in self._attributes:
                yield i
            else:
                o = {}
                isMultilineText = False
                isInit = False
                for i in self.__read_section("//RECORD Attribute"):
                    d = i.split("\t")
                    if d[0] == "ATT":
                        if isInit:
                            if self._use_cache:
                                self._attributes.append(o)
                            yield o
                        d = i.split("\t")
                        o["name"] = h.get_quotedString(d[1])
                        if d[2] and (d[2].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[2]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[2])
                            else:
                                o["text"] = (d[2].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                if isInit:
                    yield o

    Attributes = property(
        fget=getAtt,
        doc="Attributes via iterator"
        )

    def getSys(self):
        if self._use_cache and self._systems:
            for i in self._systems:
                yield i
            else:
                o = {}
                isMultilineText = False
                isInit = False
                for i in self.__read_section("//RECORD System"):
                    d = i.split("\t")
                    if d[0] == "ATT":
                        if isInit:
                            if self._use_cache:
                                self._systems.append(o)
                            yield o
                        d = i.split("\t")
                        o["name"] = h.get_quotedString(d[1])
                        if d[2] and (d[2].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[2]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[2])
                            else:
                                o["text"] = (d[2].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                if isInit:
                    yield o

    Systems = property(
        fget=getSys,
        doc="Systems via iterator"
        )

    def getComp(self):
        if self._use_cache and self._components:
            for i in self._components:
                yield i
            else:
                o = {}
                isMultilineText = False
                isInit = False
                for i in self.__read_section("//RECORD Component"):
                    d = i.split("\t")
                    if d[0] == "ATT":
                        if isInit:
                            if self._use_cache:
                                self._components.append(o)
                            yield o
                        d = i.split("\t")
                        o["name"] = h.get_quotedString(d[1])
                        if d[2] and (d[2].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[2]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[2])
                            else:
                                o["text"] = (d[2].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["notes"] = []
                        isInit = True
                    else:
                        if isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                        elif d[1] == "MEM":
                            if len(d) > 2:
                                o["notes"].append(
                                    h.get_quotedString(d[2])
                                    )
                            else:
                                o["notes"].append("**unknown**")
                if isInit:
                    yield o

    Components = property(
        fget=getComp,
        doc="Components via iterator"
        )

    def getPar(self):
        if self._use_cache and self._parameters:
            for i in self._parameters:
                yield i
            else:
                sections = [
                    r"//RECORD Probability (q)",
                    r"//RECORD Failure rate (r)",
                    r"//RECORD Frequency (f)",
                    r"//RECORD MTTR (Tr)",
                    r"//RECORD Test interval (Ti)",
                    r"//RECORD Time to first test (Tf)",
                    r"//RECORD Mission time (Tm)",
                    r"//RECORD Beta factor",
                    r"//RECORD Gamma factor",
                    r"//RECORD Delta factor",
                    r"//RECORD Alpha2 factor",
                    r"//RECORD Alpha3 factor",
                    r"//RECORD Alpha4 factor"
                    ]
                for s in sections:
                    if self._hasXLSX:
                        wb = load_workbook(
                            filename=self.xlsxpath,
                            read_only=True
                            )
                        ws = wb["Parameter-DDP"]
                        ddp = defaultdict(list)
                        for p in ws.rows[1:]:
                            t = p[0].value
                            if t == "Probability":
                                new_t = partypes[0]
                            elif t == "Failure Rate":
                                new_t = partypes[1]
                            elif t == "Frequency":
                                new_t = partypes[2]
                            elif t == "MTTR":
                                new_t = partypes[3]
                            elif t == "Test Interval":
                                new_t = partypes[4]
                            elif t == "Time To First Test":
                                new_t = partypes[5]
                            elif t == "Mission Time":
                                new_t = partypes[6]
                            elif t == "Beta Factor":
                                new_t = partypes[7]
                            elif t == "Gamma Factor":
                                new_t = partypes[8]
                            elif t == "Delta Factor":
                                new_t = partypes[9]
                            elif t == "Alpha2 Factor":
                                new_t = partypes[11]
                            elif t == "Alpha3 Factor":
                                new_t = partypes[12]
                            elif t == "Alpha4 Factor":
                                new_t = partypes[13]
                            ddp[(new_t, p[1].value)].append(
                                (float(p[2].value), float(p[3].value))
                                )
                        else:
                            ddp = {}
                    o = {}
                    isInit = False
                    isMultilineText = False
                    for i in self.__read_section(s):
                        d = i.split("\t")
                        if d[0][0] == "P":
                            if isInit:
                                if self._use_cache:
                                    self._parameters.append(o)
                                yield o
                            o["type"] = partypes[int(d[0][1:]) - 1]
                            o["name"] = h.get_quotedString(d[1])
                            o["mean"] = h.get_float(d[2])
                            dist = int(d[3])
                            if dist == 1:
                                o["distribution"] = {
                                    "type": pardist[dist],
                                    "EF": h.get_float(d[4])
                                    }
                            if dist == 2 or dist == 3:
                                o["distribution"] = {
                                    "type": pardist[dist],
                                    "alpha": h.get_float(d[4])
                                    }
                            if dist == 4:
                                o["distribution"] = {
                                    "type": pardist[dist],
                                    "std_dev": h.get_float(d[4])
                                    }
                            if dist == 5 or dist == 6:
                                o["distribution"] = {
                                    "type": pardist[dist],
                                    "min": h.get_float(d[4]),
                                    "max": h.get_float(d[5])}
                            if dist == 7:
                                ddp_val = [
                                    {"x":x, "cdf":y}
                                        for (x, y)
                                        in ddp.get((o["type"], o["name"]), {})
                                    ]
                                o["distribution"] = {
                                    "type": pardist[dist],
                                    "ddp": ddp_val
                                    }
                            if d[6] and (d[6].strip())[-1:] != r'"':
                                isMultilineText = True
                            if d[6]:
                                if not(isMultilineText):
                                    o["text"] = h.get_quotedString(d[6])
                                else:
                                    o["text"] = (d[6].strip())[1:].strip()
                            else:
                                o["text"] = ""
                            o["notes"] = []
                            isInit = True
                        else:
                            if isMultilineText:
                                idx = i.find(r'"')
                                if idx > -1:
                                    o["text"] += i[:idx]
                                    isMultilineText = False
                                else:
                                    o["text"] += i
                            elif d[1] == "MEM":
                                if len(d) > 2:
                                    o["notes"].append(
                                        h.get_quotedString(d[2])
                                        )
                                else:
                                    o["notes"].append("**unknown**")
                    # emit last
                    if isInit:
                        yield o

    Parameters = property(
        fget=getPar,
        doc="Parameters via iterator"
        )

    def getMemo(self):
        if self._use_cache and self._memo:
            for i in self._memo:
                yield i
            else:
                o = {}
                isMultilineText = False
                isInit = False
                isBody = False
                for i in self.__read_section("//RECORD Memo"):
                    if i[:3] == "MEM":
                        if isInit:
                            if self._use_cache:
                                self._memo.append(o)
                            yield o
                        d = i.split("\t")
                        o["name"] = h.get_quotedString(d[1])
                        if d[2] and (d[2].strip())[-1:] != r'"':
                            isMultilineText = True
                        if d[2]:
                            if not(isMultilineText):
                                o["text"] = h.get_quotedString(d[2])
                            else:
                                o["text"] = (d[2].strip())[1:].strip()
                        else:
                            o["text"] = ""
                        o["note"] = ""
                        isInit = True
                    else:
                        if isBody:
                            if i.strip() == "*ENDMEMO*":
                                isBody = False
                            else:
                                o["note"] += i
                        elif i.strip() == "*BEGINMEMO*":
                            isBody = True
                        elif isMultilineText:
                            idx = i.find(r'"')
                            if idx > -1:
                                o["text"] += i[:idx]
                                isMultilineText = False
                            else:
                                o["text"] += i
                if isInit:
                    yield o

    Notes = property(
        fget=getMemo,
        doc="Notes via iterator"
        )

    def getBCSets(self):
        if self._use_cache and self._bc_sets:
            for i in self._bc_sets:
                yield i
            else:
                if self._hasXLSX:
                    wb = load_workbook(filename=self.xlsxpath, read_only=True)
                    ws = wb["BC Set"]
                    bc_text = dict((e[0].value, e[1].value) for e in ws.rows)
                    ws = wb["BCSet-Input"]
                    bc_def = defaultdict(list)
                    for e in ws.rows:
                        bc_def[e[0].value].append(
                            (h.norm_string(e[1].value), e[2].value, e[3].value)
                            )
                    for n in bc_text:
                        o = {
                            "name": n,
                            "text": bc_text[n],
                            "conditions": bc_def.get(n, [])
                            }
                        if self._use_cache:
                            self._bc_sets.append(o)
                        yield o

    BCSets = property(
        fget=getBCSets,
        doc="Boundary condition sets via iterator"
        )

    def parseAnalysisRow(self, r):
        f_1 = lambda x: x if x is not None else "**unknown**"
        f_2 = lambda x: x if x is not None else ""
        key = \
            ((h.norm_string(r[0].value)).replace("_analysis_case", ""),
            r[1].value)
        topname = {"fault_tree": 8, "sequence": 8, "consequence": 9}
        doc = {
            "type": key[0],
            "name": key[1],
            "text": r[2].value,
            "setup": {
                "mcs": f_1(r[3].value),
                "uncertainty": f_1(r[4].value),
                "importance": f_1(r[5].value),
                "time_dep": f_1(r[6].value)
                },
            "conditions": f_2(r[7].value),
            "objective": f_1(r[topname[key[0]]].value)
                }
        return (key, doc)

    def getAnalysisCases(self):
        if self._use_cache and self._analyses:
            for i in self._analyses:
                yield i
            else:
                if self._hasXLSX:
                    wb = load_workbook(filename=self.xlsxpath, read_only=True)
                    ws = wb["Analysis Case"]
                    an_def = dict(self.parseAnalysisRow(r) for r in ws.rows[1:])
                    ws = wb["ConseqACase-ET"]
                    an_et = defaultdict(list)
                    for r in ws.rows[1:]:
                        an_et[("consequence", r[0].value)].append(r[1].value)
                    ws = wb["Group Spec"]
                    an_group = defaultdict(list)
                    for r in ws.rows[1:]:
                        an_group[
                            ((
                                h.norm_string(r[1].value)).replace(
                                    "_analysis_case", ""
                                    ),
                                r[0].value
                                )
                            ].append(
                                ((
                                    h.norm_string(r[2].value)).replace(
                                    "_analysis_case", ""
                                    ),
                                r[3].value
                                )
                            )
                    for a in an_def:
                        o = an_def[a]
                        if a[0] == "consequence":
                            o["event_trees"] = an_et.get(a, {})
                        elif a[0] == "mcs" or a[0] == "analysis_case_group":
                            o["filters"] = an_group.get(a, {})
                        if self._use_cache:
                            self._analyses.append(o)
                        yield o

    AnalysisCases = property(
        fget=getAnalysisCases,
        doc="Analysis cases via iterator"
        )